"""
Executive Sales Dashboard — Live Formulas
==========================================

Builds a multi-sheet executive dashboard on top of the source workbook.
All KPIs are LIVE FORMULAS that recalculate when the user edits source data.

Sheet order (left to right):
  1. Dashboard          — single-screen executive view
  2. Per-Rep Detail     — leaderboard with all metrics
  3. Stage Funnel       — full stage breakdown + chart
  4. Pipeline Health    — at-risk deals, conversion ratios
  5. Inputs             — editable quota / target / threshold inputs
  ...source sheets preserved unchanged...

Source data layout (verified against actual file):
  Madine- Conversion : header R3, data R4-R202, totals R203
  Zarrah-Conversion  : header R2, data R3-R201, totals R202
  Home-Conversion    : header R3, data R4-R202, totals R203
  Madine- Coverage   : header R3, data R4-R202, totals R203
  Zarrah- Coverage   : header R3, data R4-R202, totals R203
  Home-Coverage      : header R3, data R4-R202, totals R203

Conversion columns: B=Industry C=Coverage D=Account E=OppType F=Area
                    G=Stage H=TCV I=GP J=ClosingDate K=Probability
                    L=ProjStart M=ProjComplete N=BOMSubmission O=Remarks
Coverage columns:   B=BDM C=Account D=OppType E=Details F=Area
                    G=TCV H=GP I=ClosingDate J=Remarks K=Probability
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import (
    DataBarRule, ColorScaleRule, CellIsRule, FormulaRule,
)
from openpyxl.chart import BarChart, PieChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
import os

SRC = "/Users/hoon/executive-dashboard/source_data.xlsx"
OUT = "/Users/hoon/executive-dashboard/executive-dashboard.xlsx"

# ───────────────────────────── BU CONFIG ─────────────────────────────────────
# 6 BU slots:
#   Slots 1-3: FIXED (existing Madine, Zarrah, Home — sheet names hardcoded
#              because they have inconsistent naming and varying header rows)
#   Slots 4-6: DYNAMIC (sheet name resolved at runtime via INDIRECT() reading
#              from Inputs!RepPrefix4/5/6. Empty slots return 0 via IFERROR.
#              Convention: '<Prefix>- Coverage' and '<Prefix>- Conversion' with
#              header on row 3, data on rows 4-202 — matching Madine pattern.)
BUS = [
    {
        "name": "Madine", "slot": 1, "dynamic": False,
        "cov":  "'Madine- Coverage'",
        "conv": "'Madine- Conversion'",
        "cov_start": 4, "cov_end": 202,
        "conv_start": 4, "conv_end": 202,
    },
    {
        "name": "Zarrah", "slot": 2, "dynamic": False,
        "cov":  "'Zarrah- Coverage'",
        "conv": "'Zarrah-Conversion'",
        "cov_start": 4, "cov_end": 202,
        "conv_start": 3, "conv_end": 201,
    },
    {
        "name": "Home", "slot": 3, "dynamic": False,
        "cov":  "'Home-Coverage'",
        "conv": "'Home-Conversion'",
        "cov_start": 4, "cov_end": 202,
        "conv_start": 4, "conv_end": 202,
    },
    {
        "name": "Slot4", "slot": 4, "dynamic": True,
        "cov_start": 4, "cov_end": 202,
        "conv_start": 4, "conv_end": 202,
    },
    {
        "name": "Slot5", "slot": 5, "dynamic": True,
        "cov_start": 4, "cov_end": 202,
        "conv_start": 4, "conv_end": 202,
    },
    {
        "name": "Slot6", "slot": 6, "dynamic": True,
        "cov_start": 4, "cov_end": 202,
        "conv_start": 4, "conv_end": 202,
    },
]

STAGES_ALL    = ["Opportunity Creation", "Scoping", "BOM Submission",
                 "Resubmission", "Negotiation", "Closed Won", "Closed Lost"]
STAGES_ACTIVE = ["Opportunity Creation", "Scoping", "BOM Submission",
                 "Resubmission", "Negotiation"]

# ───────────────────────────── DESIGN TOKENS ─────────────────────────────────
NAVY        = "1A2B4A"
STEEL       = "2C4A72"
ACCENT      = "1E6FBF"
KPI_BG      = "0F1F38"
GRAY_900    = "111827"
GRAY_700    = "374151"
GRAY_500    = "6B7280"
GRAY_200    = "E5E7EB"
GRAY_100    = "F3F4F6"
WHITE       = "FFFFFF"
GREEN       = "166534"
GREEN_LITE  = "DCFCE7"
AMBER       = "92400E"
AMBER_LITE  = "FEF3C7"
RED         = "991B1B"
RED_LITE    = "FEE2E2"
MUTED_BLUE  = "9BB5D4"

PESO  = '"₱"#,##0;-"₱"#,##0;"₱"-;@'
# PESO_M used to be abbreviated millions; now identical to PESO per user request — full digits everywhere
PESO_M = PESO
PCT   = "0.0%"
INT   = "#,##0"

# ───────────────────────────── HELPERS ───────────────────────────────────────
def fill(c):
    return PatternFill("solid", fgColor=c)

def font(size=9, bold=False, color=GRAY_900, italic=False):
    return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)

ALIGN_C  = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_L  = Alignment(horizontal="left",   vertical="center", wrap_text=False, indent=1)
ALIGN_R  = Alignment(horizontal="right",  vertical="center", wrap_text=False, indent=1)
ALIGN_CW = Alignment(horizontal="center", vertical="center", wrap_text=True)

def cell(ws, r, c, v=None, f=None, fc=None, a=None, nf=None, b=None):
    x = ws.cell(row=r, column=c)
    if v is not None: x.value = v
    if f is not None: x.font = f
    if fc is not None: x.fill = fill(fc)
    if a is not None: x.alignment = a
    if nf is not None: x.number_format = nf
    if b is not None: x.border = b
    return x

def merge(ws, r1, c1, r2, c2, v=None, f=None, fc=None, a=None, nf=None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    if fc is not None:
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                ws.cell(r, c).fill = fill(fc)
    return cell(ws, r1, c1, v=v, f=f, a=a, nf=nf)

def fill_range(ws, r1, c1, r2, c2, color):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(r, c).fill = fill(color)

# ───────────────────────────── FORMULA BUILDERS ──────────────────────────────
def conv_rng(bu, col):
    """Return a conv-sheet range. For dynamic slots, returns INDIRECT(...) that
    resolves at runtime from Inputs!RepPrefix{slot}. For fixed slots, returns
    a direct ref like 'Madine- Conversion'!G4:G202."""
    if bu.get("dynamic"):
        return (f'INDIRECT("\'"&RepPrefix{bu["slot"]}&'
                f'"- Conversion\'!{col}{bu["conv_start"]}:{col}{bu["conv_end"]}")')
    return f"{bu['conv']}!{col}{bu['conv_start']}:{col}{bu['conv_end']}"

def cov_rng(bu, col):
    if bu.get("dynamic"):
        return (f'INDIRECT("\'"&RepPrefix{bu["slot"]}&'
                f'"- Coverage\'!{col}{bu["cov_start"]}:{col}{bu["cov_end"]}")')
    return f"{bu['cov']}!{col}{bu['cov_start']}:{col}{bu['cov_end']}"

def _wrap_iferror(bu, expr, default="0"):
    """Dynamic slots wrap with IF(prefix="", 0, IFERROR(expr, 0)).
    The IF short-circuits empty prefixes so INDIRECT never runs in that case
    (more robust across Excel/Numbers/LibreOffice). IFERROR catches any
    INDIRECT failures when the prefix is set but sheet doesn't exist."""
    if bu.get("dynamic"):
        return f'IF(RepPrefix{bu["slot"]}="",{default},IFERROR({expr},{default}))'
    return expr

def f_sumifs_stage(bu, sum_col, stage):
    """SUMIFS on conv: sum sum_col where stage = exact stage."""
    expr = (f'SUMIFS({conv_rng(bu, sum_col)},'
            f'{conv_rng(bu, "G")},"{stage}")')
    return _wrap_iferror(bu, expr)

def f_countif_stage(bu, stage):
    expr = f'COUNTIF({conv_rng(bu, "G")},"{stage}")'
    return _wrap_iferror(bu, expr)

def f_sum_active_tcv(bu, col="H"):
    """Sum TCV (or any col) where stage is in ACTIVE list (excludes Won/Lost/blank).
    Single SUMIFS using exclusion criteria — way shorter than per-stage expansion."""
    expr = (f'SUMIFS({conv_rng(bu, col)},'
            f'{conv_rng(bu, "G")},"<>Closed Lost",'
            f'{conv_rng(bu, "G")},"<>Closed Won",'
            f'{conv_rng(bu, "G")},"<>")')
    return _wrap_iferror(bu, expr)

def f_sum_pipeline_tcv(bu, col="H"):
    """Sum TCV across all non-Lost stages (active + Closed Won)."""
    expr = (f'SUMIFS({conv_rng(bu, col)},'
            f'{conv_rng(bu, "G")},"<>Closed Lost",'
            f'{conv_rng(bu, "G")},"<>")')
    return _wrap_iferror(bu, expr)

def f_weighted_pipeline(bu):
    """Weighted active pipeline: SUM(TCV * stage_probability) for active stages.
    SUMIFS can't multiply across ranges so per-stage. Wrap each in IFERROR
    only for dynamic slots."""
    probs = {
        "Opportunity Creation": 0.1, "Scoping": 0.3, "BOM Submission": 0.5,
        "Resubmission": 0.7, "Negotiation": 0.9,
    }
    parts = [f"{f_sumifs_stage(bu, 'H', s)}*{p}" for s, p in probs.items()]
    return "(" + "+".join(parts) + ")"

def f_count_active(bu):
    """Count active opps (anything not Won/Lost/blank)."""
    expr = (f'COUNTIFS({conv_rng(bu, "G")},"<>Closed Lost",'
            f'{conv_rng(bu, "G")},"<>Closed Won",'
            f'{conv_rng(bu, "G")},"<>")')
    return _wrap_iferror(bu, expr)

def f_count_won(bu):
    return f_countif_stage(bu, "Closed Won")

def f_count_lost(bu):
    return f_countif_stage(bu, "Closed Lost")

def f_sum_won(bu, col="H"):
    return f_sumifs_stage(bu, col, "Closed Won")

def f_sum_lost(bu, col="H"):
    return f_sumifs_stage(bu, col, "Closed Lost")

def f_count_coverage(bu):
    """Count coverage leads = rows where account name is non-empty."""
    expr = f'COUNTIFS({cov_rng(bu, "C")},"<>")'
    return _wrap_iferror(bu, expr)

def f_past_due(bu):
    """Count of opportunities past closing date in active stages.
    Uses ISNUMBER for date validity (works in Excel/Numbers/LibreOffice)."""
    j = conv_rng(bu, "J")
    g = conv_rng(bu, "G")
    parts = [f'SUMPRODUCT(--ISNUMBER({j}),--({j}<TODAY()),--({g}="{s}"))' for s in STAGES_ACTIVE]
    expr = "(" + "+".join(parts) + ")"
    return _wrap_iferror(bu, expr)

def f_past_due_tcv(bu):
    j = conv_rng(bu, "J")
    g = conv_rng(bu, "G")
    h = conv_rng(bu, "H")
    parts = [f'SUMPRODUCT(--ISNUMBER({j}),--({j}<TODAY()),--({g}="{s}"),{h})' for s in STAGES_ACTIVE]
    expr = "(" + "+".join(parts) + ")"
    return _wrap_iferror(bu, expr)

def sum_across_bus(per_bu_formula):
    """Wrap a per-BU formula generator into a sum across all BUs."""
    return "+".join(per_bu_formula(bu) for bu in BUS)

# ═════════════════════════ MAIN BUILD ════════════════════════════════════════
print(f"[1/7] Loading source workbook: {SRC}")
wb = openpyxl.load_workbook(SRC)

# ─────────────────────────── HELPER COLUMN ───────────────────────────────────
# Add hidden column T to each fixed conv sheet with "Eligible TCV":
# returns the TCV when (stage<>"Closed Lost" AND stage<>"" AND ISNUMBER(tcv)),
# else "". This sidesteps the LARGE(IF(...)) array-formula CSE requirement
# that breaks in iOS/iPad mobile Excel — LARGE on a flat numeric column
# works in EVERY version of Excel (2010+, web, mobile, mac, desktop).
def add_helper_col(sheet_name, header_row, data_start, data_end):
    ws_src = wb[sheet_name]
    ws_src.cell(header_row, 20).value = "Eligible TCV (helper)"
    ws_src.cell(header_row, 20).font = Font(name="Calibri", size=8,
                                             italic=True, color="999999")
    for r in range(data_start, data_end + 1):
        ws_src.cell(r, 20).value = (
            f'=IF(OR(G{r}="Closed Lost",G{r}="",NOT(ISNUMBER(H{r}))),"",H{r})')
    ws_src.column_dimensions['T'].hidden = True

add_helper_col('Madine- Conversion', 3, 4, 202)
add_helper_col('Zarrah-Conversion',  2, 3, 201)
add_helper_col('Home-Conversion',    3, 4, 202)

# Remove any prior dashboard sheets we'll regenerate
for sn in ("Dashboard", "Per-Rep Detail", "Stage Funnel", "Pipeline Health", "Inputs"):
    if sn in wb.sheetnames:
        del wb[sn]

# Build sheets in REVERSE order so they end up at the FRONT (insert at index 0)
print("[2/7] Creating Inputs sheet")
wb_inputs = wb.create_sheet("Inputs", 0)

print("[3/7] Creating Pipeline Health sheet")
wb_health = wb.create_sheet("Pipeline Health", 0)

print("[4/7] Creating Stage Funnel sheet")
wb_funnel = wb.create_sheet("Stage Funnel", 0)

print("[5/7] Creating Per-Rep Detail sheet")
wb_rep = wb.create_sheet("Per-Rep Detail", 0)

print("[6/7] Creating Dashboard sheet (main)")
ws = wb.create_sheet("Dashboard", 0)

# ════════════════════════════════════════════════════════════════════════════
# SHEET: Inputs (editable quota, targets)
# ════════════════════════════════════════════════════════════════════════════
def build_inputs(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 50

    merge(ws, 1, 1, 1, 4, fc=NAVY)
    merge(ws, 2, 1, 2, 4, v="DASHBOARD INPUTS · EDIT VALUES IN COLUMN C",
          f=font(14, True, WHITE), fc=NAVY, a=ALIGN_L)
    merge(ws, 3, 1, 3, 4, fc=NAVY)
    ws.row_dimensions[1].height = 6
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 6

    # ── Block A: Targets/Thresholds (rows 5-13) ─────────────────────────────
    rows = [
        ("Quarterly TCV Quota (₱)",        500_000_000, "Used to compute Pipeline Coverage Ratio.", PESO),
        ("Pipeline Coverage Target (×)",   3.0,         "Healthy B2B benchmark: 3-5× quota.",      "0.0\"×\""),
        ("Target GP Margin (%)",           0.20,        "Threshold for green/amber on GP%.",       PCT),
        ("Stage-Aging Threshold (days)",   60,          "Deals older than this in same stage = at-risk.", INT),
        ("At-Risk Deal Size Threshold (₱)", 100_000_000, "Past-due deals above this auto-flag red.", PESO),
    ]
    for i, (label, value, note, nf) in enumerate(rows):
        r = 5 + i * 2
        cell(ws, r, 2, label, f=font(10, True), fc=GRAY_100, a=ALIGN_L)
        cell(ws, r, 3, value, f=font(11, True, ACCENT), fc=WHITE, a=ALIGN_R, nf=nf)
        cell(ws, r, 4, note, f=font(9, color=GRAY_500, italic=True), fc=GRAY_100, a=ALIGN_L)
        ws.row_dimensions[r].height = 26
        ws.row_dimensions[r + 1].height = 4

    # ── Block B: Branding (rows 15-17) ──────────────────────────────────────
    branding = [
        ("Dashboard Title",    "EXECUTIVE SALES DASHBOARD",            "Edit to rebrand the title strip.", "@"),
        ("Dashboard Subtitle", "Live pipeline · TCV · GP · Stage · Health", "Subtitle below the title.",  "@"),
    ]
    for i, (label, value, note, nf) in enumerate(branding):
        r = 15 + i * 2
        cell(ws, r, 2, label, f=font(10, True), fc=GRAY_100, a=ALIGN_L)
        cell(ws, r, 3, value, f=font(11, True, ACCENT), fc=WHITE, a=ALIGN_R, nf=nf)
        cell(ws, r, 4, note, f=font(9, color=GRAY_500, italic=True), fc=GRAY_100, a=ALIGN_L)
        ws.row_dimensions[r].height = 26
        ws.row_dimensions[r + 1].height = 4

    # ── Block C: Rep Slots header (row 20) ──────────────────────────────────
    ws.row_dimensions[20].height = 28
    merge(ws, 20, 2, 20, 4, v="REP SLOTS — Set Sheet Prefix to wire each slot to a Coverage/Conversion sheet pair",
          fc=NAVY, f=font(10, True, WHITE), a=ALIGN_L)

    # ── Block D: 6 Rep Slots (rows 22-44, alternating prefix/display/spacer) ─
    # Layout per slot: Sheet Prefix row, Display Name row, spacer
    rep_slots = [
        (1, "Madine", "Madine", "Slot 1 — pre-wired to existing Madine sheets (do not change prefix)."),
        (2, "Zarrah", "Zarrah", "Slot 2 — pre-wired to existing Zarrah sheets (do not change prefix)."),
        (3, "Home",   "Home",   "Slot 3 — pre-wired to existing Home sheets (do not change prefix)."),
        (4, "",       "Rep 4",  "Slot 4 — empty. To activate: create '<Prefix>- Coverage' & '<Prefix>- Conversion' sheets, then type prefix here."),
        (5, "",       "Rep 5",  "Slot 5 — empty. Same activation steps as Slot 4."),
        (6, "",       "Rep 6",  "Slot 6 — empty. Same activation steps as Slot 4."),
    ]
    for i, (slot, prefix, display, note) in enumerate(rep_slots):
        base = 22 + i * 3
        # Prefix row
        cell(ws, base, 2,     f"Rep {slot} Sheet Prefix", f=font(10, True), fc=GRAY_100, a=ALIGN_L)
        cell(ws, base, 3,     prefix, f=font(11, True, ACCENT), fc=WHITE, a=ALIGN_R, nf="@")
        cell(ws, base, 4,     note,   f=font(9, color=GRAY_500, italic=True), fc=GRAY_100, a=ALIGN_L)
        ws.row_dimensions[base].height = 26
        # Display name row
        cell(ws, base+1, 2,   f"Rep {slot} Display Name", f=font(10, True), fc=GRAY_100, a=ALIGN_L)
        cell(ws, base+1, 3,   display, f=font(11, True, ACCENT), fc=WHITE, a=ALIGN_R, nf="@")
        cell(ws, base+1, 4,   "Shown on dashboard. Independent of sheet prefix.",
             f=font(9, color=GRAY_500, italic=True), fc=GRAY_100, a=ALIGN_L)
        ws.row_dimensions[base+1].height = 22
        # Spacer
        ws.row_dimensions[base+2].height = 4

    # ── Block E: How-to-add-a-rep instructions (rows 42+) ───────────────────
    instr_start = 42
    ws.row_dimensions[instr_start].height = 22
    merge(ws, instr_start, 2, instr_start, 4,
          v="HOW TO ADD A NEW REP (Slots 4–6)",
          fc=NAVY, f=font(11, True, WHITE), a=ALIGN_L)
    instructions = [
        "1. Right-click the 'Madine- Coverage' tab → 'Move or Copy' → check 'Create a copy' → OK.",
        "2. Right-click the new tab → Rename to '<NewName>- Coverage' (must include '- ' with space).",
        "3. Open the renamed sheet → clear all data rows below the header (keep row 1-3 untouched).",
        "4. Repeat steps 1–3 for 'Madine- Conversion' → '<NewName>- Conversion'.",
        "5. Type the new rep's data into both sheets.",
        "6. Come back to this Inputs sheet → set 'Rep 4 Sheet Prefix' (cell C31) to '<NewName>'.",
        "7. Optional: set 'Rep 4 Display Name' (cell C32) to a friendlier label.",
        "8. The Dashboard automatically includes the new rep — no rebuild needed.",
        "",
        "✓ The prefix MUST exactly match the sheet name minus '- Coverage' / '- Conversion'.",
        "✓ Slots 1-3 are PRE-WIRED to Madine / Zarrah / Home — leave their prefixes alone.",
        "✓ To remove a slot, simply blank out its prefix → totals exclude it automatically.",
        "✓ A hidden helper column 'T' on each conversion sheet auto-tracks eligible deals — it",
        "    copies along when you duplicate the Madine sheet, so don't worry about it.",
    ]
    for i, line in enumerate(instructions):
        r = instr_start + 1 + i
        ws.row_dimensions[r].height = 18
        cell(ws, r, 2, "" if line == "" else line,
             f=font(9, color=GRAY_900), fc=WHITE, a=ALIGN_L)

    # Named cells — used throughout the dashboard for editable references
    from openpyxl.workbook.defined_name import DefinedName
    name_defs = [
        ("Quota",      "Inputs!$C$5"),
        ("CovTarget",  "Inputs!$C$7"),
        ("GPTarget",   "Inputs!$C$9"),
        ("AgeThresh",  "Inputs!$C$11"),
        ("RiskThresh", "Inputs!$C$13"),
        ("DashTitle",  "Inputs!$C$15"),
        ("DashSub",    "Inputs!$C$17"),
        # Slot 1 = rows 22 (prefix), 23 (display name)
        ("RepPrefix1", "Inputs!$C$22"),
        ("RepName1",   "Inputs!$C$23"),
        ("RepPrefix2", "Inputs!$C$25"),
        ("RepName2",   "Inputs!$C$26"),
        ("RepPrefix3", "Inputs!$C$28"),
        ("RepName3",   "Inputs!$C$29"),
        ("RepPrefix4", "Inputs!$C$31"),
        ("RepName4",   "Inputs!$C$32"),
        ("RepPrefix5", "Inputs!$C$34"),
        ("RepName5",   "Inputs!$C$35"),
        ("RepPrefix6", "Inputs!$C$37"),
        ("RepName6",   "Inputs!$C$38"),
    ]
    for name, ref in name_defs:
        wb.defined_names[name] = DefinedName(name, attr_text=ref)

build_inputs(wb_inputs)

# ════════════════════════════════════════════════════════════════════════════
# SHEET: Dashboard (main exec view)
# ════════════════════════════════════════════════════════════════════════════
def build_dashboard(ws):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B6"

    # Column widths — sized for full-digit currency (e.g. ₱3,893,192,450 = 14 chars)
    widths = {1: 2.2, 2: 14, 3: 17, 4: 17, 5: 11, 6: 17, 7: 2.5,
              8: 19, 9: 8, 10: 17, 11: 17, 12: 9, 13: 2.5, 14: 14}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    # ── ROW 1-4 TITLE STRIP ─────────────────────────────────────────────────
    ws.row_dimensions[1].height = 6
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 8

    fill_range(ws, 1, 1, 4, 14, NAVY)
    merge(ws, 2, 2, 2, 7, v="=DashTitle",
          f=font(15, True, WHITE), a=ALIGN_L)
    merge(ws, 3, 2, 3, 7, v="=DashSub",
          f=font(9, color=MUTED_BLUE, italic=True), a=ALIGN_L)
    merge(ws, 2, 8, 2, 14,
          v='=UPPER(TEXT(TODAY(),"DDDD, DD MMMM YYYY"))',
          f=font(11, True, WHITE), a=ALIGN_R)
    merge(ws, 3, 8, 3, 14,
          v='=RepName1&"  ·  "&RepName2&"  ·  "&RepName3',
          f=font(9, color=MUTED_BLUE, italic=True), a=ALIGN_R)

    # ── ROW 5-6 KPI TILES ────────────────────────────────────────────────────
    ws.row_dimensions[5].height = 46
    ws.row_dimensions[6].height = 18

    # KPI tiles reference Per-Rep Detail's TOTAL row (row 11). This is more
    # compat-friendly than chaining 6 BU formulas: SUM gracefully handles
    # empty slots, and the Per-Rep total row already does the heavy lifting.
    f_total_tcv = "='Per-Rep Detail'!H11"
    f_total_gp  = "='Per-Rep Detail'!I11"
    f_gp_margin = "=IFERROR('Per-Rep Detail'!I11/'Per-Rep Detail'!H11,0)"
    f_weighted  = "='Per-Rep Detail'!K11"
    f_active    = "='Per-Rep Detail'!D11"
    f_winrate   = "='Per-Rep Detail'!G11"

    tiles = [
        (2, 3,  "TOTAL TCV",       f_total_tcv,  PESO_M),
        (4, 5,  "TOTAL GP",        f_total_gp,   PESO_M),
        (6, 6,  "GP MARGIN",       f_gp_margin,  PCT),
        (8, 9,  "WEIGHTED PIPE",   f_weighted,   PESO_M),
        (10, 11,"ACTIVE OPPS",     f_active,     INT),
        (12, 14,"WIN RATE",        f_winrate,    PCT),
    ]
    for c1, c2, label, formula, nf in tiles:
        merge(ws, 5, c1, 5, c2, v=formula,
              f=font(18, True, WHITE), fc=KPI_BG, a=ALIGN_C, nf=nf)
        merge(ws, 6, c1, 6, c2, v=label,
              f=font(8, True, MUTED_BLUE), fc=STEEL, a=ALIGN_C)

    # Fill gutters
    for r in (5, 6):
        for c in (1, 7, 13):
            cell(ws, r, c, fc=NAVY)

    # ── ROW 7-8 SECTION SPACER + HEADERS ─────────────────────────────────────
    ws.row_dimensions[7].height = 8
    ws.row_dimensions[8].height = 18
    fill_range(ws, 7, 1, 7, 14, NAVY)

    merge(ws, 8, 2, 8, 6,  v="REP LEADERBOARD",  f=font(9, True, MUTED_BLUE), fc=NAVY, a=ALIGN_L)
    merge(ws, 8, 8, 8, 14, v="STAGE BREAKDOWN",  f=font(9, True, MUTED_BLUE), fc=NAVY, a=ALIGN_L)
    cell(ws, 8, 1, fc=NAVY); cell(ws, 8, 7, fc=NAVY); cell(ws, 8, 13, fc=NAVY)

    # ── ROW 9 HEADERS ────────────────────────────────────────────────────────
    ws.row_dimensions[9].height = 22
    lb_headers = [(2, "REP"), (3, "TCV"), (4, "GP"),
                  (5, "GP%"), (6, "WTD PIPE")]
    for c, h in lb_headers:
        cell(ws, 9, c, h, f=font(8, True, WHITE), fc=STEEL, a=ALIGN_C)

    fn_headers = [(8, "STAGE"), (9, "OPPS"), (10, "TCV"),
                  (11, "WTD"), (12, "PROB")]
    for c, h in fn_headers:
        cell(ws, 9, c, h, f=font(8, True, WHITE), fc=STEEL, a=ALIGN_C)
    cell(ws, 9, 14, "Δ vs Quota", f=font(8, True, WHITE), fc=STEEL, a=ALIGN_C)

    cell(ws, 9, 1, fc=NAVY); cell(ws, 9, 7, fc=NAVY); cell(ws, 9, 13, fc=NAVY)

    # ── ROWS 10-15 LEADERBOARD (6 rep slots, row 16 = TOTAL) ────────────────
    # All 6 BUs shown. Slots 4-6 will be blank rows when their prefix is empty
    # (display name formula returns "" and IFERROR catches the value formulas).
    # Empty slots silently contribute 0 to totals via IFERROR-wrapped formulas.
    rep_layout = [
        # (row, BUS index, name_ref, bg, slot_num)
        (10, 2, "=RepName3", GRAY_100, 3),  # Home
        (11, 1, "=RepName2", WHITE,    2),  # Zarrah
        (12, 0, "=RepName1", GRAY_100, 1),  # Madine
        (13, 3, "=RepName4", WHITE,    4),  # Slot 4 (dynamic)
        (14, 4, "=RepName5", GRAY_100, 5),  # Slot 5 (dynamic)
        (15, 5, "=RepName6", WHITE,    6),  # Slot 6 (dynamic)
    ]
    for r, bu_idx, name_ref, bg, slot in rep_layout:
        bu = BUS[bu_idx]
        ws.row_dimensions[r].height = 22

        # For dynamic slots, blank the display name when prefix is empty.
        # For fixed slots, show the display name (RepName1/2/3) directly.
        if bu.get("dynamic"):
            display_formula = f'=IF(RepPrefix{slot}="","",{name_ref[1:]})'
        else:
            display_formula = name_ref
        cell(ws, r, 2, display_formula, f=font(10, True, GRAY_900), fc=bg, a=ALIGN_L)

        # Numeric cells: blank for dynamic slots when prefix empty, otherwise compute.
        def slot_aware(formula):
            if bu.get("dynamic"):
                return f'=IF(RepPrefix{slot}="","",{formula})'
            return f"={formula}"

        # TCV (pipeline = active + won, excludes Closed Lost)
        cell(ws, r, 3, slot_aware(f"({f_sum_pipeline_tcv(bu, 'H')})"),
             f=font(10, color=GRAY_900), fc=bg, a=ALIGN_R, nf=PESO_M)
        # GP
        cell(ws, r, 4, slot_aware(f"({f_sum_pipeline_tcv(bu, 'I')})"),
             f=font(10, color=GRAY_900), fc=bg, a=ALIGN_R, nf=PESO_M)
        # GP%
        cell(ws, r, 5, slot_aware(
                f"IFERROR(({f_sum_pipeline_tcv(bu, 'I')})/({f_sum_pipeline_tcv(bu, 'H')}),0)"),
             f=font(10, True, GRAY_900), fc=bg, a=ALIGN_C, nf=PCT)
        # Weighted Pipeline
        cell(ws, r, 6, slot_aware(f"({f_weighted_pipeline(bu)})"),
             f=font(10, color=GRAY_900), fc=bg, a=ALIGN_R, nf=PESO_M)
        cell(ws, r, 1, fc=NAVY); cell(ws, r, 7, fc=NAVY); cell(ws, r, 13, fc=NAVY)

    # Total row 16
    r = 16
    ws.row_dimensions[r].height = 22
    cell(ws, r, 2, "TOTAL", f=font(10, True, WHITE), fc=STEEL, a=ALIGN_L)
    cell(ws, r, 3, "=SUM(C10:C15)", f=font(10, True, WHITE), fc=STEEL, a=ALIGN_R, nf=PESO_M)
    cell(ws, r, 4, "=SUM(D10:D15)", f=font(10, True, WHITE), fc=STEEL, a=ALIGN_R, nf=PESO_M)
    cell(ws, r, 5, "=IFERROR(D16/C16,0)", f=font(10, True, WHITE), fc=STEEL, a=ALIGN_C, nf=PCT)
    cell(ws, r, 6, "=SUM(F10:F15)", f=font(10, True, WHITE), fc=STEEL, a=ALIGN_R, nf=PESO_M)
    cell(ws, r, 1, fc=NAVY); cell(ws, r, 7, fc=NAVY); cell(ws, r, 13, fc=NAVY)

    # ── ROWS 10-15 STAGE BREAKDOWN (right side, cols H:N) ────────────────────
    # Active stages + Closed Won + Closed Lost
    stage_data = [
        ("Opportunity Creation", 0.1, GRAY_100),
        ("Scoping",              0.3, WHITE),
        ("BOM Submission",       0.5, GRAY_100),
        ("Resubmission",         0.7, WHITE),
        ("Negotiation",          0.9, GRAY_100),
        ("Closed Won",           1.0, GREEN_LITE),
    ]
    for i, (stage, prob, bg) in enumerate(stage_data):
        r = 10 + i
        if r > 15:
            break  # only have rows 10-15 for stage section here (6 stages)
        ws.row_dimensions[r].height = 22
        # Stage label
        cell(ws, r, 8, stage, f=font(9, True, GRAY_900), fc=bg, a=ALIGN_L)
        # Count
        cnt_f = "=" + sum_across_bus(lambda b: f_countif_stage(b, stage))
        cell(ws, r, 9, cnt_f, f=font(10, True, GRAY_900), fc=bg, a=ALIGN_C, nf=INT)
        # TCV
        tcv_f = "=" + sum_across_bus(lambda b: f_sumifs_stage(b, "H", stage))
        cell(ws, r, 10, tcv_f, f=font(10, color=GRAY_900), fc=bg, a=ALIGN_R, nf=PESO_M)
        # Weighted
        cell(ws, r, 11, f"=J{r}*{prob}", f=font(10, color=GRAY_900), fc=bg, a=ALIGN_R, nf=PESO_M)
        # Probability
        cell(ws, r, 12, prob, f=font(10, True, GRAY_900), fc=bg, a=ALIGN_C, nf=PCT)
        cell(ws, r, 1, fc=NAVY); cell(ws, r, 7, fc=NAVY); cell(ws, r, 13, fc=NAVY)
        cell(ws, r, 14, fc=GRAY_100)

    # Pipeline coverage on N10 (uses F16 = new leaderboard TOTAL row).
    cell(ws, 10, 14,
         "=IFERROR(F16/Quota,0)",
         f=font(11, True, ACCENT), fc=GRAY_100, a=ALIGN_C, nf='0.0"×"')
    cell(ws, 11, 14, "Pipe÷Quota", f=font(7, color=GRAY_500), fc=GRAY_100, a=ALIGN_C)
    fill_range(ws, 12, 14, 13, 14, GRAY_100)

    # ── ROW 16 CLOSED LOST + ROW 17 PIPELINE TOTAL ───────────────────────────
    # Stage rows 10-15 have 6 stages; row 16 = Closed Lost; row 17 = TOTAL
    ws.row_dimensions[16].height = 22
    cell(ws, 16, 8, "Closed Lost", f=font(9, True, GRAY_900), fc=RED_LITE, a=ALIGN_L)
    cell(ws, 16, 9, "=" + sum_across_bus(lambda b: f_countif_stage(b, "Closed Lost")),
         f=font(10, True, GRAY_900), fc=RED_LITE, a=ALIGN_C, nf=INT)
    cell(ws, 16, 10, "=" + sum_across_bus(lambda b: f_sumifs_stage(b, "H", "Closed Lost")),
         f=font(10, color=GRAY_900), fc=RED_LITE, a=ALIGN_R, nf=PESO_M)
    cell(ws, 16, 11, 0, f=font(10, color=GRAY_900), fc=RED_LITE, a=ALIGN_R, nf=PESO_M)
    cell(ws, 16, 12, 0, f=font(10, True, GRAY_900), fc=RED_LITE, a=ALIGN_C, nf=PCT)
    cell(ws, 16, 1, fc=NAVY); cell(ws, 16, 7, fc=NAVY); cell(ws, 16, 13, fc=NAVY); cell(ws, 16, 14, fc=GRAY_100)

    # Stage TOTAL row 17 (pipeline excludes Closed Lost = SUM rows 10-15)
    ws.row_dimensions[17].height = 22
    r = 17
    cell(ws, r, 8, "PIPELINE TOTAL", f=font(9, True, WHITE), fc=STEEL, a=ALIGN_L)
    cell(ws, r, 9, "=SUM(I10:I15)", f=font(10, True, WHITE), fc=STEEL, a=ALIGN_C, nf=INT)
    cell(ws, r, 10, "=SUM(J10:J15)", f=font(10, True, WHITE), fc=STEEL, a=ALIGN_R, nf=PESO_M)
    cell(ws, r, 11, "=SUM(K10:K15)", f=font(10, True, WHITE), fc=STEEL, a=ALIGN_R, nf=PESO_M)
    cell(ws, r, 12, "", fc=STEEL)
    cell(ws, r, 1, fc=NAVY); cell(ws, r, 7, fc=NAVY); cell(ws, r, 13, fc=NAVY); cell(ws, r, 14, fc=GRAY_100)

    # Conditional formatting on leaderboard
    ws.conditional_formatting.add(
        "C10:C12",
        DataBarRule(start_type="num", start_value=0, end_type="max", color=ACCENT)
    )
    ws.conditional_formatting.add(
        "F10:F12",
        DataBarRule(start_type="num", start_value=0, end_type="max", color="60A5FA")
    )
    ws.conditional_formatting.add(
        "E10:E12",
        ColorScaleRule(
            start_type="num", start_value=0.10, start_color=RED_LITE,
            mid_type="num",   mid_value=0.20,   mid_color=AMBER_LITE,
            end_type="num",   end_value=0.30,   end_color=GREEN_LITE,
        )
    )

    # ── ROW 18 SPACER + HEALTH HEADER ────────────────────────────────────────
    ws.row_dimensions[18].height = 8
    ws.row_dimensions[19].height = 18
    fill_range(ws, 18, 1, 18, 14, NAVY)
    merge(ws, 19, 2, 19, 14,
          v="PIPELINE HEALTH · AT-RISK & PAST-DUE DEALS",
          f=font(9, True, MUTED_BLUE), fc=NAVY, a=ALIGN_L)
    cell(ws, 19, 1, fc=NAVY)

    # ── ROW 20-21 HEALTH KPI MINI-TILES ──────────────────────────────────────
    ws.row_dimensions[20].height = 38
    ws.row_dimensions[21].height = 16

    # Health tiles also reference Per-Rep Detail aggregates where possible.
    # Past-Due Count: SUM of per-rep past-due column (M5:M10).
    # Coverage Leads: SUM of per-rep coverage column (C5:C10).
    # Conv Rate: (active + won + lost) / coverage = D11+E11+F11 / C11
    f_pastdue_count = "=SUM('Per-Rep Detail'!M5:M10)"
    f_pastdue_tcv   = "=" + sum_across_bus(f_past_due_tcv)
    f_cov_total     = "=SUM('Per-Rep Detail'!C5:C10)"
    f_conv_rate     = ("=IFERROR(('Per-Rep Detail'!D11+'Per-Rep Detail'!E11+'Per-Rep Detail'!F11)"
                       "/'Per-Rep Detail'!C11,0)")
    # Lost Rate by ₱: closed lost TCV / (closed won + closed lost) TCV — keep direct
    f_lost_rate     = (f"=IFERROR(({sum_across_bus(lambda b: f_sum_lost(b, 'H'))})/"
                       f"((({sum_across_bus(lambda b: f_sum_won(b, 'H'))}))+"
                       f"({sum_across_bus(lambda b: f_sum_lost(b, 'H'))})),0)")

    health_tiles = [
        (2, 3, "PAST-DUE DEALS",  f_pastdue_count, INT),
        (4, 5, "PAST-DUE TCV",    f_pastdue_tcv,   PESO_M),
        (6, 7, "COVERAGE LEADS",  f_cov_total,     INT),
        (8, 9, "COV→CONV %",      f_conv_rate,     PCT),
        (10, 11, "LOST RATE (₱)", f_lost_rate,     PCT),
        (12, 14, "PIPE COVERAGE", "=IFERROR(F16/Quota,0)", '0.0"× quota"'),
    ]
    for c1, c2, label, formula, nf in health_tiles:
        merge(ws, 20, c1, 20, c2, v=formula,
              f=font(15, True, WHITE), fc=KPI_BG, a=ALIGN_C, nf=nf)
        merge(ws, 21, c1, 21, c2, v=label,
              f=font(7, True, MUTED_BLUE), fc=STEEL, a=ALIGN_C)
    cell(ws, 20, 1, fc=NAVY); cell(ws, 21, 1, fc=NAVY)

    # ── ROW 22 SPACER + TOP DEALS HEADER ─────────────────────────────────────
    ws.row_dimensions[22].height = 8
    ws.row_dimensions[23].height = 18
    fill_range(ws, 22, 1, 22, 14, NAVY)
    merge(ws, 23, 2, 23, 14,
          v="TOP DEALS — DRILL TO 'Pipeline Health' SHEET FOR FULL LIST",
          f=font(9, True, MUTED_BLUE), fc=NAVY, a=ALIGN_L)
    cell(ws, 23, 1, fc=NAVY)

    # Headers row 24
    ws.row_dimensions[24].height = 22
    merge(ws, 24, 2, 24, 4, v="ACCOUNT", f=font(8, True, WHITE), fc=STEEL, a=ALIGN_L)
    cell(ws,  24, 5, "REP",      f=font(8, True, WHITE), fc=STEEL, a=ALIGN_C)
    merge(ws, 24, 6, 24, 7, v="STAGE",   f=font(8, True, WHITE), fc=STEEL, a=ALIGN_L)
    merge(ws, 24, 8, 24, 9, v="TCV",     f=font(8, True, WHITE), fc=STEEL, a=ALIGN_R)
    merge(ws, 24, 10, 24, 11, v="GP",    f=font(8, True, WHITE), fc=STEEL, a=ALIGN_R)
    merge(ws, 24, 12, 24, 13, v="CLOSING", f=font(8, True, WHITE), fc=STEEL, a=ALIGN_C)
    cell(ws, 24, 14, "FLAG", f=font(8, True, WHITE), fc=STEEL, a=ALIGN_C)
    cell(ws, 24, 1, fc=NAVY)

    # Top 6 deals — formulas link to Pipeline Health rows 5-10 (top of Home top-8)
    for i in range(25, 31):
        ws.row_dimensions[i].height = 18
        cell(ws, i, 1, fc=NAVY)
        ph_row = i - 25 + 5  # Pipeline Health row (5,6,7,8,9,10)
        bg = GRAY_100 if i % 2 == 1 else WHITE
        merge(ws, i, 2, i, 4, v=f"='Pipeline Health'!B{ph_row}",
              f=font(9, color=GRAY_900), fc=bg, a=ALIGN_L)
        cell(ws, i, 5, f"='Pipeline Health'!C{ph_row}",
             f=font(9, color=GRAY_900), fc=bg, a=ALIGN_C)
        merge(ws, i, 6, i, 7, v=f"='Pipeline Health'!D{ph_row}",
              f=font(9, color=GRAY_900), fc=bg, a=ALIGN_L)
        merge(ws, i, 8, i, 9, v=f"='Pipeline Health'!E{ph_row}",
              f=font(9, color=GRAY_900), fc=bg, a=ALIGN_R, nf=PESO_M)
        merge(ws, i, 10, i, 11, v=f"='Pipeline Health'!F{ph_row}",
              f=font(9, color=GRAY_900), fc=bg, a=ALIGN_R, nf=PESO_M)
        merge(ws, i, 12, i, 13, v=f"='Pipeline Health'!G{ph_row}",
              f=font(9, color=GRAY_900), fc=bg, a=ALIGN_C, nf="MMM DD")
        cell(ws, i, 14, f"='Pipeline Health'!H{ph_row}",
             f=font(9, True, GRAY_900), fc=bg, a=ALIGN_C)

    # ── ROW 31 FOOTER ────────────────────────────────────────────────────────
    ws.row_dimensions[31].height = 8
    fill_range(ws, 31, 1, 31, 14, NAVY)
    ws.row_dimensions[32].height = 16
    merge(ws, 32, 2, 32, 14,
          v="Live formulas · Edit source sheets to refresh · Inputs sheet drives quota/thresholds",
          f=font(8, color=MUTED_BLUE, italic=True), fc=NAVY, a=ALIGN_L)
    cell(ws, 32, 1, fc=NAVY)
    ws.row_dimensions[33].height = 8
    fill_range(ws, 33, 1, 33, 14, NAVY)

build_dashboard(ws)

# ════════════════════════════════════════════════════════════════════════════
# SHEET: Per-Rep Detail — full leaderboard with all metrics
# ════════════════════════════════════════════════════════════════════════════
def build_per_rep(ws):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B5"

    # Per-Rep Detail: widen TCV/GP/Wtd Pipe columns for full-digit values
    widths = [3, 14, 14, 12, 11, 11, 11, 18, 17, 11, 18, 17, 13, 3]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # Title strip
    ws.row_dimensions[1].height = 6
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 8
    fill_range(ws, 1, 1, 3, 14, NAVY)
    merge(ws, 2, 2, 2, 13, v="PER-REP DETAIL · LEADERBOARD",
          f=font(14, True, WHITE), a=ALIGN_L)

    # Headers row 4
    ws.row_dimensions[4].height = 36
    headers = [
        (2,  "REP",            ALIGN_L),
        (3,  "Coverage Leads", ALIGN_C),
        (4,  "Active Opps",    ALIGN_C),
        (5,  "Closed Won",     ALIGN_C),
        (6,  "Closed Lost",    ALIGN_C),
        (7,  "Win Rate",       ALIGN_C),
        (8,  "TCV (Pipe+Won)", ALIGN_R),
        (9,  "GP (Pipe+Won)",  ALIGN_R),
        (10, "GP %",           ALIGN_C),
        (11, "Weighted Pipe",  ALIGN_R),
        (12, "Avg Deal Size",  ALIGN_R),
        (13, "Past-Due Deals", ALIGN_C),
    ]
    for c, h, a in headers:
        cell(ws, 4, c, h, f=font(8, True, WHITE), fc=STEEL, a=a)
    cell(ws, 4, 1, fc=NAVY); cell(ws, 4, 14, fc=NAVY)

    # 6 rep slots — slots 4-6 are dynamic (empty when prefix unset)
    rep_order = [
        ("Madine", BUS[0], "=RepName1", 1),
        ("Zarrah", BUS[1], "=RepName2", 2),
        ("Home",   BUS[2], "=RepName3", 3),
        ("Slot4",  BUS[3], "=RepName4", 4),
        ("Slot5",  BUS[4], "=RepName5", 5),
        ("Slot6",  BUS[5], "=RepName6", 6),
    ]
    for i, (name, bu, name_ref, slot) in enumerate(rep_order):
        r = 5 + i
        ws.row_dimensions[r].height = 24
        bg = WHITE if i % 2 == 0 else GRAY_100

        # Display name: blank when slot empty
        if bu.get("dynamic"):
            disp = f'=IF(RepPrefix{slot}="","",{name_ref[1:]})'
        else:
            disp = name_ref
        cell(ws, r, 2, disp, f=font(11, True, GRAY_900), fc=bg, a=ALIGN_L)

        def slot_val(formula):
            if bu.get("dynamic"):
                return f'=IF(RepPrefix{slot}="","",{formula})'
            return f"={formula}"

        cell(ws, r, 3, slot_val(f_count_coverage(bu)),
             f=font(10), fc=bg, a=ALIGN_C, nf=INT)
        cell(ws, r, 4, slot_val(f"({f_count_active(bu)})"),
             f=font(10), fc=bg, a=ALIGN_C, nf=INT)
        cell(ws, r, 5, slot_val(f_count_won(bu)),
             f=font(10), fc=bg, a=ALIGN_C, nf=INT)
        cell(ws, r, 6, slot_val(f_count_lost(bu)),
             f=font(10), fc=bg, a=ALIGN_C, nf=INT)
        cell(ws, r, 7, slot_val(
                f"IFERROR({f_count_won(bu)}/({f_count_won(bu)}+{f_count_lost(bu)}),0)"),
             f=font(10, True), fc=bg, a=ALIGN_C, nf=PCT)
        cell(ws, r, 8, slot_val(f"({f_sum_pipeline_tcv(bu, 'H')})"),
             f=font(10), fc=bg, a=ALIGN_R, nf=PESO_M)
        cell(ws, r, 9, slot_val(f"({f_sum_pipeline_tcv(bu, 'I')})"),
             f=font(10), fc=bg, a=ALIGN_R, nf=PESO_M)
        cell(ws, r, 10, f"=IFERROR(I{r}/H{r},0)",
             f=font(10, True), fc=bg, a=ALIGN_C, nf=PCT)
        cell(ws, r, 11, slot_val(f"({f_weighted_pipeline(bu)})"),
             f=font(10), fc=bg, a=ALIGN_R, nf=PESO_M)
        cell(ws, r, 12, f"=IFERROR(H{r}/(D{r}+E{r}),0)",
             f=font(10), fc=bg, a=ALIGN_R, nf=PESO_M)
        cell(ws, r, 13, slot_val(f"({f_past_due(bu)})"),
             f=font(10, True), fc=bg, a=ALIGN_C, nf=INT)
        cell(ws, r, 1, fc=NAVY); cell(ws, r, 14, fc=NAVY)

    # Total row (rep rows occupy 5-10, total on row 11)
    r = 11
    ws.row_dimensions[r].height = 26
    cell(ws, r, 2, "TOTAL", f=font(11, True, WHITE), fc=STEEL, a=ALIGN_L)
    for c in range(3, 14):
        if c == 7:  # Win Rate — recompute won/(won+lost) from sums of the 6 rows
            cell(ws, r, c,
                 f"=IFERROR(SUM(E5:E10)/(SUM(E5:E10)+SUM(F5:F10)),0)",
                 f=font(11, True, WHITE), fc=STEEL, a=ALIGN_C, nf=PCT)
        elif c == 10:
            cell(ws, r, c, f"=IFERROR(I{r}/H{r},0)",
                 f=font(11, True, WHITE), fc=STEEL, a=ALIGN_C, nf=PCT)
        elif c == 12:
            cell(ws, r, c, f"=IFERROR(H{r}/(D{r}+E{r}),0)",
                 f=font(11, True, WHITE), fc=STEEL, a=ALIGN_R, nf=PESO_M)
        else:
            col = get_column_letter(c)
            nf = INT if c in (3, 4, 5, 6, 13) else PESO_M
            align = ALIGN_C if c in (3, 4, 5, 6, 13) else ALIGN_R
            cell(ws, r, c, f"=SUM({col}5:{col}10)",
                 f=font(11, True, WHITE), fc=STEEL, a=align, nf=nf)
    cell(ws, r, 1, fc=NAVY); cell(ws, r, 14, fc=NAVY)

    # Conditional formatting (now spans 6 rep rows: 5-10)
    ws.conditional_formatting.add("H5:H10",
        DataBarRule(start_type="num", start_value=0, end_type="max", color=ACCENT))
    ws.conditional_formatting.add("K5:K10",
        DataBarRule(start_type="num", start_value=0, end_type="max", color="60A5FA"))
    ws.conditional_formatting.add("J5:J10",
        ColorScaleRule(start_type="num", start_value=0.10, start_color=RED_LITE,
                       mid_type="num", mid_value=0.20, mid_color=AMBER_LITE,
                       end_type="num", end_value=0.30, end_color=GREEN_LITE))
    ws.conditional_formatting.add("M5:M10",
        CellIsRule(operator="greaterThan", formula=["0"], fill=fill(RED_LITE),
                   font=font(10, True, RED)))

    # Rep-level chart: TCV by rep (now 6 rows)
    chart = BarChart()
    chart.type = "bar"
    chart.style = 11
    chart.title = "Pipeline TCV by Rep"
    chart.y_axis.title = ""
    chart.x_axis.title = ""
    data = Reference(ws, min_col=8, min_row=4, max_row=10, max_col=8)
    cats = Reference(ws, min_col=2, min_row=5, max_row=10)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 18
    ws.add_chart(chart, "B14")

    # Weighted pipeline chart
    chart2 = BarChart()
    chart2.type = "bar"
    chart2.style = 12
    chart2.title = "Weighted Pipeline by Rep"
    data2 = Reference(ws, min_col=11, min_row=4, max_row=10, max_col=11)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats)
    chart2.height = 8
    chart2.width = 18
    ws.add_chart(chart2, "H14")

build_per_rep(wb_rep)

# ════════════════════════════════════════════════════════════════════════════
# SHEET: Stage Funnel — per-stage breakdown across all BUs
# ════════════════════════════════════════════════════════════════════════════
def build_funnel(ws):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B5"

    # Stage Funnel: 6 BU count columns + Total + TCV + Weighted
    widths = [3, 22, 11, 11, 11, 11, 11, 11, 11, 11, 18, 17, 3]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # Title strip
    ws.row_dimensions[1].height = 6
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 8
    fill_range(ws, 1, 1, 3, 13, NAVY)
    merge(ws, 2, 2, 2, 12, v="STAGE FUNNEL · PIPELINE BY STAGE",
          f=font(14, True, WHITE), a=ALIGN_L)

    # Header row 4 — 6 rep count columns (4-9), Total Opps (10), TCV (11), Wtd (12)
    ws.row_dimensions[4].height = 32
    cell(ws, 4, 2,  "STAGE",       f=font(9, True, WHITE), fc=STEEL, a=ALIGN_C)
    cell(ws, 4, 3,  "Probability", f=font(9, True, WHITE), fc=STEEL, a=ALIGN_C)
    # 6 BU columns — display names from Inputs (slots 4-6 blank when prefix empty)
    for slot in range(1, 7):
        if slot <= 3:
            cell(ws, 4, 3 + slot, f'=RepName{slot}&" #"',
                 f=font(9, True, WHITE), fc=STEEL, a=ALIGN_C)
        else:
            cell(ws, 4, 3 + slot, f'=IF(RepPrefix{slot}="","—",RepName{slot}&" #")',
                 f=font(9, True, WHITE), fc=STEEL, a=ALIGN_C)
    cell(ws, 4, 10, "Total Opps", f=font(9, True, WHITE), fc=STEEL, a=ALIGN_C)
    cell(ws, 4, 11, "Total TCV",  f=font(9, True, WHITE), fc=STEEL, a=ALIGN_C)
    cell(ws, 4, 12, "Weighted",   f=font(9, True, WHITE), fc=STEEL, a=ALIGN_C)
    cell(ws, 4, 1, fc=NAVY); cell(ws, 4, 13, fc=NAVY)

    # Stage rows (active stages → won → lost)
    stages = [
        ("Opportunity Creation", 0.1),
        ("Scoping",              0.3),
        ("BOM Submission",       0.5),
        ("Resubmission",         0.7),
        ("Negotiation",          0.9),
        ("Closed Won",           1.0),
        ("Closed Lost",          0.0),
    ]
    for i, (stage, prob) in enumerate(stages):
        r = 5 + i
        ws.row_dimensions[r].height = 22
        bg = WHITE if i % 2 == 0 else GRAY_100
        if stage == "Closed Won":
            bg = GREEN_LITE
        elif stage == "Closed Lost":
            bg = RED_LITE

        cell(ws, r, 2, stage, f=font(10, True, GRAY_900), fc=bg, a=ALIGN_L)
        cell(ws, r, 3, prob,  f=font(10), fc=bg, a=ALIGN_C, nf=PCT)
        # Per BU counts (6 columns: cols 4-9)
        for slot, col in zip(range(1, 7), [4, 5, 6, 7, 8, 9]):
            bu = BUS[slot - 1]
            if bu.get("dynamic"):
                f = f'=IF(RepPrefix{slot}="","",{f_countif_stage(bu, stage)})'
            else:
                f = f'={f_countif_stage(bu, stage)}'
            cell(ws, r, col, f, f=font(10), fc=bg, a=ALIGN_C, nf=INT)
        # Total opps (col 10)
        cell(ws, r, 10, f"=SUM(D{r}:I{r})",
             f=font(10, True), fc=bg, a=ALIGN_C, nf=INT)
        # Total TCV (col 11)
        cell(ws, r, 11,
             "=" + sum_across_bus(lambda b: f_sumifs_stage(b, "H", stage)),
             f=font(10), fc=bg, a=ALIGN_R, nf=PESO_M)
        # Weighted (col 12)
        cell(ws, r, 12, f"=K{r}*C{r}", f=font(10), fc=bg, a=ALIGN_R, nf=PESO_M)
        cell(ws, r, 1, fc=NAVY); cell(ws, r, 13, fc=NAVY)

    # Total row (rows 5-11 are 7 stages, total on row 12)
    r = 12
    ws.row_dimensions[r].height = 26
    cell(ws, r, 2, "PIPELINE TOTAL (excl. Lost)", f=font(11, True, WHITE), fc=STEEL, a=ALIGN_L)
    cell(ws, r, 3, "", fc=STEEL)
    # 6 BU count cols (4-9) + Total Opps (10) — sum rows 5-10 to exclude Closed Lost row 11
    for c in range(4, 11):
        col = get_column_letter(c)
        cell(ws, r, c, f"=SUM({col}5:{col}10)",
             f=font(11, True, WHITE), fc=STEEL, a=ALIGN_C, nf=INT)
    cell(ws, r, 11, "=SUM(K5:K10)", f=font(11, True, WHITE), fc=STEEL, a=ALIGN_R, nf=PESO_M)
    cell(ws, r, 12, "=SUM(L5:L10)", f=font(11, True, WHITE), fc=STEEL, a=ALIGN_R, nf=PESO_M)
    cell(ws, r, 1, fc=NAVY); cell(ws, r, 13, fc=NAVY)

    # Bar chart: opps per stage (Total Opps column = 10)
    chart = BarChart()
    chart.type = "bar"
    chart.style = 11
    chart.title = "Opps Count by Stage"
    data = Reference(ws, min_col=10, min_row=4, max_row=11, max_col=10)
    cats = Reference(ws, min_col=2, min_row=5, max_row=11)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 10
    chart.width = 18
    ws.add_chart(chart, "B14")

    # Doughnut: TCV by stage (Total TCV column = 11; excl Lost row)
    chart2 = DoughnutChart()
    chart2.title = "TCV Distribution by Stage"
    data2 = Reference(ws, min_col=11, min_row=5, max_row=10)
    cats2 = Reference(ws, min_col=2, min_row=5, max_row=10)
    chart2.add_data(data2, titles_from_data=False)
    chart2.set_categories(cats2)
    chart2.height = 10
    chart2.width = 11
    ws.add_chart(chart2, "H14")

build_funnel(wb_funnel)

# ════════════════════════════════════════════════════════════════════════════
# SHEET: Pipeline Health — top deals + at-risk
# ════════════════════════════════════════════════════════════════════════════
def build_health(ws):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B5"

    # Pipeline Health: widen account name + numeric columns
    widths = [3, 32, 10, 17, 17, 17, 13, 16, 30, 3]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # Title
    ws.row_dimensions[1].height = 6
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 8
    fill_range(ws, 1, 1, 3, 10, NAVY)
    merge(ws, 2, 2, 2, 9, v="PIPELINE HEALTH · TOP DEALS & AT-RISK",
          f=font(14, True, WHITE), a=ALIGN_L)

    # Header
    ws.row_dimensions[4].height = 28
    headers = [(2, "ACCOUNT"), (3, "REP"), (4, "STAGE"),
               (5, "TCV"), (6, "GP"), (7, "CLOSING"),
               (8, "FLAG"), (9, "NOTES")]
    for c, h in headers:
        cell(ws, 4, c, h, f=font(9, True, WHITE), fc=STEEL, a=ALIGN_C)
    cell(ws, 4, 1, fc=NAVY); cell(ws, 4, 10, fc=NAVY)

    # Top 24 deals — pulled live via LARGE/INDEX/MATCH from all 3 conv sheets
    # We use a helper trick: create an aggregated view via formulas that
    # reference each BU's conv sheet by row index.
    # Simpler approach: pre-write top-deals from all 3 sheets sorted by TCV.
    # Use Excel formula array to find top 24 across all 3 sheets.
    # Since openpyxl doesn't easily do array formulas, we'll write per-row
    # formulas pulling the top deal (by TCV) from each sheet, alternating.
    # Concretely: row 5-9 = top 5 Home, row 10-14 = top 5 Zarrah, row 15-19 = top 5 Madine,
    # then sorted view at rows 20+ via LARGE-based formulas.

    # Top-N using hidden helper column T (Eligible TCV) on each conv sheet.
    # LARGE on a flat numeric column needs no array context and no CSE — works
    # in iOS/Android mobile Excel, Excel for the Web, Mac Excel, and desktop alike.
    # MATCH lookups go against col T (filtered values) so they line up with
    # the LARGE result; INDEX then pulls the matching account/stage/etc.
    rows_per_bu = 8
    cur_row = 5
    for name, bu, name_ref, slot in [
        ("Home",   BUS[2], "=RepName3", 3),
        ("Zarrah", BUS[1], "=RepName2", 2),
        ("Madine", BUS[0], "=RepName1", 1),
        ("Slot4",  BUS[3], "=RepName4", 4),
        ("Slot5",  BUS[4], "=RepName5", 5),
        ("Slot6",  BUS[5], "=RepName6", 6),
    ]:
        for rank in range(1, rows_per_bu + 1):
            r = cur_row
            cur_row += 1
            ws.row_dimensions[r].height = 18
            bg = WHITE if (r - 5) % 2 == 0 else GRAY_100

            tcv_helper  = conv_rng(bu, "T")  # eligible-TCV helper column
            acct_range  = conv_rng(bu, "D")
            stage_range = conv_rng(bu, "G")
            close_range = conv_rng(bu, "J")
            gp_range    = conv_rng(bu, "I")

            large_tcv = f'LARGE({tcv_helper},{rank})'
            tcv_f   = f'=IFERROR({large_tcv},"")'
            acct_f  = f'=IFERROR(INDEX({acct_range},MATCH({large_tcv},{tcv_helper},0)),"")'
            stage_f = f'=IFERROR(INDEX({stage_range},MATCH({large_tcv},{tcv_helper},0)),"")'
            close_f = f'=IFERROR(INDEX({close_range},MATCH({large_tcv},{tcv_helper},0)),"")'
            gp_f    = f'=IFERROR(INDEX({gp_range},MATCH({large_tcv},{tcv_helper},0)),"")'

            # Rep label: blank when row has no account (rank exceeds deal count
            # OR dynamic slot is unused). Account formula at B{r} drives everything.
            inner_name = name_ref[1:]  # strips '=' prefix
            if bu.get("dynamic"):
                rep_display = (f'=IF(OR(RepPrefix{slot}="",B{r}=""),"",'
                               f'{inner_name})')
            else:
                rep_display = f'=IF(B{r}="","",{inner_name})'

            cell(ws, r, 2, acct_f, f=font(9, color=GRAY_900), fc=bg, a=ALIGN_L)
            cell(ws, r, 3, rep_display, f=font(9, True, GRAY_900), fc=bg, a=ALIGN_C)
            cell(ws, r, 4, stage_f, f=font(9, color=GRAY_900), fc=bg, a=ALIGN_L)
            cell(ws, r, 5, tcv_f, f=font(9, color=GRAY_900), fc=bg, a=ALIGN_R, nf=PESO_M)
            cell(ws, r, 6, gp_f, f=font(9, color=GRAY_900), fc=bg, a=ALIGN_R, nf=PESO_M)
            cell(ws, r, 7, close_f, f=font(9, color=GRAY_900), fc=bg, a=ALIGN_C, nf="MMM DD")

            # Flag formula: blank if no account (empty rank or unused slot),
            # else: WON / LOST / PAST DUE / NO DATE / NEGOTIATING / ACTIVE
            flag_f = (f'=IF(B{r}="","",'
                      f'IF(D{r}="Closed Won","✓ WON",'
                      f'IF(D{r}="Closed Lost","✗ LOST",'
                      f'IF(AND(G{r}<>"",ISNUMBER(G{r}),G{r}<TODAY()),"⚠ PAST DUE",'
                      f'IF(G{r}="","◔ NO DATE",'
                      f'IF(D{r}="Negotiation","● NEGOTIATING","○ ACTIVE"))))))')
            cell(ws, r, 8, flag_f, f=font(9, True, GRAY_900), fc=bg, a=ALIGN_C)
            cell(ws, r, 9, "", fc=bg)
            cell(ws, r, 1, fc=NAVY); cell(ws, r, 10, fc=NAVY)

    # Conditional formatting on Flag column
    # Past Due → red fill
    health_end = cur_row - 1
    rng = f"H5:H{health_end}"
    ws.conditional_formatting.add(rng,
        FormulaRule(formula=[f'ISNUMBER(SEARCH("PAST DUE",H5))'],
                    fill=fill(RED_LITE), font=font(9, True, RED)))
    ws.conditional_formatting.add(rng,
        FormulaRule(formula=[f'ISNUMBER(SEARCH("WON",H5))'],
                    fill=fill(GREEN_LITE), font=font(9, True, GREEN)))
    ws.conditional_formatting.add(rng,
        FormulaRule(formula=[f'ISNUMBER(SEARCH("LOST",H5))'],
                    fill=fill(GRAY_200), font=font(9, color=GRAY_700)))
    ws.conditional_formatting.add(rng,
        FormulaRule(formula=[f'ISNUMBER(SEARCH("NO DATE",H5))'],
                    fill=fill(AMBER_LITE), font=font(9, True, AMBER)))
    # Data bars on TCV
    ws.conditional_formatting.add(f"E5:E{health_end}",
        DataBarRule(start_type="num", start_value=0, end_type="max", color=ACCENT))

build_health(wb_health)

# ════════════════════════════════════════════════════════════════════════════
# Re-order sheets: Dashboard, Per-Rep, Stage Funnel, Pipeline Health, Inputs, then sources
# ════════════════════════════════════════════════════════════════════════════
desired_order = ["Dashboard", "Per-Rep Detail", "Stage Funnel", "Pipeline Health", "Inputs"]
existing = wb.sheetnames
new_order = desired_order + [s for s in existing if s not in desired_order]
# Build by re-indexing
wb._sheets = [wb[s] for s in new_order]

# Set Dashboard as the active sheet
wb.active = wb.sheetnames.index("Dashboard")

print("[7/7] Saving:", OUT)
wb.save(OUT)
print(f"✓ Done. File size: {os.path.getsize(OUT):,} bytes")
